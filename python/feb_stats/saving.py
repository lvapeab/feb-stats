from io import BytesIO
from typing import Optional

import pandas as pd
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from python.feb_stats.entities import League
from python.feb_stats.entities_ops import average_games
from python.feb_stats.lang import spanish_columns
from python.feb_stats.utils import get_sorted_list_of_columns, timedelta_to_str


def league_to_xlsx(
        league: League,
        filename: Optional[str] = None,
        col_width: int = 30,
        export_language: str = "es",
) -> bytes:
    """Exports a league to xlsx.
    :param league: League to be exported.
    :param filename: Name of the file to be written.
    :param col_width: Column width.
    :param export_language: Export the league in different languages. Currently, only 'es' or 'en' supported.
    :return: The exported xlsx file, as bytes.
    """

    if league.aggregated_games is None:
        raise ValueError(f'The league {league} has no aggregated games.')
    filename = filename or f'{league.name}_{league.season.replace("/", "-")}.xlsx'

    xlsx_writer = pd.ExcelWriter(
        filename, engine="openpyxl", mode="w", date_format="DD-MM-YYYY"
    )

    sheet_name = f'{league.name}_{league.season.replace("/", "-")}'
    columns = get_sorted_list_of_columns()
    column_names = (
        map(lambda x: spanish_columns[x], columns)
        if export_language == "es"
        else columns
    )
    column_names = list(
        map(lambda x: " ".join(x.capitalize().split("_")), column_names)
    )

    aggregated_games = league.aggregated_games
    averaged_games = average_games(aggregated_games.copy())

    aggregated_games.loc[:, "minutes"] = aggregated_games["minutes"].apply(
        lambda x: timedelta_to_str(x) if not pd.isnull(x) else ""
    )
    averaged_games.loc[:, "minutes"] = averaged_games["minutes"].apply(
        lambda x: timedelta_to_str(x) if not pd.isnull(x) else ""
    )
    aggregated_games.to_excel(
        xlsx_writer,
        float_format="%.3f",
        columns=columns,
        encoding="latin1",
        header=column_names,
        sheet_name=sheet_name,
    )

    averaged_games.to_excel(
        xlsx_writer,
        float_format="%.3f",
        columns=columns,
        encoding="latin1",
        header=column_names,
        sheet_name=sheet_name + "-medias",
    )

    player_columns = get_sorted_list_of_columns(individual_columns=True)
    player_column_names = (
        map(lambda x: spanish_columns[x], player_columns)
        if export_language == "es"
        else player_columns
    )
    player_column_names = list(
        map(lambda x: " ".join(x.capitalize().split("_")), player_column_names)
    )
    for team in league.teams:
        if team.season_stats is not None:
            aggregated_team_season_games = team.season_stats
            averaged_team_season_games = average_games(
                aggregated_team_season_games.copy(), individual_columns=True
            )
            aggregated_team_season_games.loc[:, "minutes"] = \
                aggregated_team_season_games["minutes"].apply(lambda x: timedelta_to_str(x) if not pd.isnull(x) else "")
            averaged_team_season_games.loc[:, "minutes"] = \
                averaged_team_season_games["minutes"].apply(lambda x: timedelta_to_str(x) if not pd.isnull(x) else "")
            aggregated_team_season_games.to_excel(
                xlsx_writer,
                float_format="%.3f",
                columns=player_columns,
                header=player_column_names,
                encoding="latin1",
                sheet_name=team.name[:31],
            )
            averaged_team_season_games.to_excel(
                xlsx_writer,
                float_format="%.3f",
                columns=player_columns,
                header=player_column_names,
                encoding="latin1",
                sheet_name=f"medias - {team.name}"[:31],
            )

    center_alignment = Alignment(horizontal="center")
    for n_sheet, (worksheet_name, worksheet) in enumerate(xlsx_writer.sheets.items()):
        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = center_alignment
        for n_column in range(worksheet.max_column):
            worksheet.column_dimensions[get_column_letter(n_column + 1)].width = (
                2 * col_width if n_column < 2 else col_width
            )
    virtual_workbook = BytesIO()
    xlsx_writer.book.save(virtual_workbook)
    virtual_workbook.seek(0)
    return virtual_workbook.read()
