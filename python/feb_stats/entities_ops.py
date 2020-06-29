from io import BytesIO
import pandas as pd
import numpy as np
from typing import TypeVar, Generic, List, Optional
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

from python.feb_stats.entities import League, Team, Game, Boxscore
from python.feb_stats.utils import get_averageable_numerical_columns, get_sorted_list_of_columns, timedelta_to_str


# TODO: These ops should be done in a DB
def get_team_by_name(league: League,
                     team_name: str) -> Team:
    for team in league.teams:
        if team.name == team_name:
            return team
    raise Exception(f'Unable to find the team {team_name} in the league {league}')


def get_games_by_team(league: League,
                      team: Team) -> List[Game]:
    matching_games = []
    for game in league.games:
        if team in {game.local_team, game.away_team}:
            matching_games.append(game)
    return matching_games


def get_team_boxscores(league: League,
                       team: Team) -> List[Boxscore]:
    return [game.local_boxscore if game.local_team == team else game.away_boxscore
            for game in get_games_by_team(league, team)]


def get_rival_boxscores(league: League,
                        team: Team) -> List[Boxscore]:
    return [game.local_boxscore if game.local_team != team else game.away_boxscore
            for game in get_games_by_team(league, team)]


def average_games(df: pd.DataFrame,
                  individual_columns: bool = False) -> pd.DataFrame:
    n_games = df.loc[:, 'partidos'].astype(np.float32)
    df.loc[:, get_averageable_numerical_columns(individual_columns=individual_columns)] = \
        df.loc[:, get_averageable_numerical_columns(individual_columns=individual_columns)].astype(np.float32).div(n_games, axis='rows')
    if 'minutos' in df:
        df.loc[:, 'minutos'] /= n_games
    df.loc[:, 'modo'] = 'Media'
    return df


def league_to_excel(league,
                    filename: str = None,
                    col_width: int = 30) -> bytes:
    filename = filename or f'{league.name}_{league.season.replace("/", "-")}.xlsx'

    # Create a Pandas Excel writer using openpyxl as the engine.
    writer = pd.ExcelWriter(filename,
                            engine='openpyxl',
                            mode='w',
                            date_format='DD-MM-YYYY'
                            )

    if league.aggregated_games is not None:
        sheet_name = f'{league.name}_{league.season.replace("/", "-")}'
        columns = get_sorted_list_of_columns()
        column_names = list(map(lambda x: ' '.join(x.capitalize().split('_')),
                                columns))

        aggregated_games = league.aggregated_games
        averaged_games = average_games(aggregated_games.copy())

        aggregated_games.loc[:, 'minutos'] = aggregated_games['minutos'].apply(
            lambda x: timedelta_to_str(x) if not pd.isnull(x) else ''
        )
        averaged_games.loc[:, 'minutos'] = averaged_games['minutos'].apply(
            lambda x: timedelta_to_str(x) if not pd.isnull(x) else ''
        )
        aggregated_games.to_excel(
            writer,
            float_format="%.3f",
            columns=columns,
            encoding='latin1',
            header=column_names,
            sheet_name=sheet_name)

        averaged_games.to_excel(
            writer,
            float_format="%.3f",
            columns=columns,
            encoding='latin1',
            header=column_names,
            sheet_name=sheet_name + '-medias')

        player_columns = get_sorted_list_of_columns(individual_columns=True)
        player_column_names = list(map(lambda x: ' '.join(x.capitalize().split('_')),
                                       player_columns))
        for team in league.teams:
            if team.season_stats is not None:
                aggregated_team_season_games = team.season_stats
                averaged_team_season_games = average_games(aggregated_team_season_games.copy(),
                                                           individual_columns=True
                                                           )

                aggregated_team_season_games.loc[:, 'minutos'] = aggregated_team_season_games['minutos'].apply(
                    lambda x: timedelta_to_str(x) if not pd.isnull(x) else ''
                )
                averaged_team_season_games.loc[:, 'minutos'] = averaged_team_season_games['minutos'].apply(
                    lambda x: timedelta_to_str(x) if not pd.isnull(x) else ''
                )
                aggregated_team_season_games.to_excel(
                    writer,
                    float_format="%.3f",
                    columns=player_columns,
                    header=player_column_names,
                    encoding='latin1',
                    sheet_name=team.name[:31])
                averaged_team_season_games.to_excel(
                    writer,
                    float_format="%.3f",
                    columns=player_columns,
                    header=player_column_names,
                    encoding='latin1',
                    sheet_name=f'medias - {team.name}'[:31])


        center_alignment = Alignment(horizontal='center')
        for n_sheet, (worksheet_name, worksheet) in enumerate(writer.sheets.items()):
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.alignment = center_alignment

            for n_column in range(worksheet.max_column):
                worksheet.column_dimensions[get_column_letter(n_column + 1)].width = 2 * col_width if n_column < 2 else col_width
        virtual_workbook = BytesIO()
        writer.book.save(virtual_workbook)
        virtual_workbook.seek(0)
        return virtual_workbook.read()
