from io import BytesIO

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from src.core.analysis.entities import League
from src.core.analysis.entities_ops import average_games
from src.core.analysis.lang import spanish_columns
from src.core.analysis.utils import get_sorted_list_of_columns, timedelta_to_str

VERY_HIGH_VALUES_COLOR = "green"
HIGH_VALUES_COLOR = "yellow"
LOW_VALUES_COLOR = "orange"
VERY_LOW_VALUES_COLOR = "red"


def gaussian_color_style(
    s: pd.Series,
    very_high_regime: str = f"background-color: {VERY_HIGH_VALUES_COLOR}; border: 1px solid",
    high_regime: str = f"background-color: {HIGH_VALUES_COLOR}; border: 1px solid",
    low_regime: str = f"background-color: {LOW_VALUES_COLOR}; border: 1px solid",
    very_low_regime: str = f"background-color: {VERY_LOW_VALUES_COLOR}; border: 1px solid",
    regular_regime: str = "border: 1px solid",
) -> list[str]:
    high_indices = np.where(s > s.mean() + s.std(), high_regime, regular_regime)
    very_high_indices = np.where(s > s.mean() + 2 * s.std(), very_high_regime, regular_regime)
    low_indices = np.where(s < s.mean() - s.std(), low_regime, regular_regime)
    very_low_indices = np.where(s < s.mean() - 2 * s.std(), very_low_regime, regular_regime)
    out_color = []
    for i in range(len(high_indices)):
        if very_high_indices[i] != regular_regime:
            out_color.append(very_high_indices[i])
            continue
        if very_low_indices[i] != regular_regime:
            out_color.append(very_low_indices[i])
            continue
        if high_indices[i] != regular_regime:
            out_color.append(high_indices[i])
            continue
        out_color.append(low_indices[i])
    return out_color


def format_minutes_column(df: pd.DataFrame) -> None:
    """Format timedelta minutes column to strings."""
    if "minutes" not in df.columns:
        return

    formatted_minutes = df["minutes"].map(lambda x: (timedelta_to_str(x, minute_format="02d")))
    df["minutes"] = df["minutes"].astype("object")
    df["minutes"] = formatted_minutes
    return


def league_to_xlsx(
    league: League,
    filename: str | None = None,
    col_width: int = 60,
    export_language: str = "es",
    export_colors: bool = False,
) -> bytes:
    """Exports a league to xlsx.
    :param league: League to be exported.
    :param filename: Name of the file to be written.
    :param col_width: Column width.
    :param export_language: Export the league in different languages. Currently, only 'es' or 'en' supported.
    :param export_colors: Add a color style to the output xlsx file.
    :return: The exported xlsx file, as bytes.
    """

    if league.aggregated_games is None:
        raise ValueError(f"The league {league} has no aggregated games.")
    filename = filename or f'{league.name}_{league.season.replace("/", "-")}.xlsx'

    xlsx_writer = pd.ExcelWriter(filename, engine="openpyxl", mode="w", date_format="DD-MM-YYYY")

    sheet_name = f'{league.name}_{league.season.replace("/", "-")}'
    columns = get_sorted_list_of_columns()

    aggregated_games = league.aggregated_games.loc[:, columns]
    aggregated_games = aggregated_games.set_index("team")
    columns.pop(columns.index("team"))
    averaged_games = average_games(aggregated_games.copy())

    aggregated_games.loc[:, "minutes"] = aggregated_games["minutes"].apply(
        lambda x: timedelta_to_str(x) if not pd.isnull(x) else ""
    )
    averaged_games.loc[:, "minutes"] = averaged_games["minutes"].apply(
        lambda x: timedelta_to_str(x) if not pd.isnull(x) else ""
    )
    numerical_columns = list(
        set(aggregated_games.columns)
        - {
            "mode",
            "minutes",
        }
    )
    if export_colors:
        aggregated_games = aggregated_games.style.apply(
            gaussian_color_style,
            subset=numerical_columns,
            axis=0,
        )

        averaged_games = averaged_games.style.apply(
            gaussian_color_style,
            subset=numerical_columns,
            axis=0,
        )
    column_names = list(map(lambda x: spanish_columns[x], columns) if export_language == "es" else columns)
    column_names = list(map(lambda x: "\n".join(x.split()), column_names))

    aggregated_games.to_excel(
        xlsx_writer,
        float_format="%.2f",
        columns=columns,
        header=column_names,
        sheet_name=sheet_name,
    )

    averaged_games.to_excel(
        xlsx_writer,
        float_format="%.2f",
        columns=columns,
        header=column_names,
        sheet_name=f"Medias {sheet_name}",
    )

    player_columns = get_sorted_list_of_columns(individual_columns=True)
    player_column_names = list(
        map(lambda x: spanish_columns[x], player_columns) if export_language == "es" else player_columns
    )
    player_column_names = list(map(lambda x: "\n".join(x.split()), player_column_names))
    for team in league.teams:
        if team.season_stats is not None:
            aggregated_team_season_games = team.season_stats.loc[:, player_columns]
            averaged_team_season_games = average_games(aggregated_team_season_games.copy(), individual_columns=True)
            format_minutes_column(aggregated_team_season_games)
            format_minutes_column(averaged_team_season_games)
            numerical_columns = list(
                set(aggregated_team_season_games.columns) - {"mode", "minutes", "player", "number"}
            )
            if export_colors:
                aggregated_team_season_games = aggregated_team_season_games.style.apply(
                    gaussian_color_style,
                    subset=numerical_columns,
                    axis=0,
                )

                averaged_team_season_games = averaged_team_season_games.style.apply(
                    gaussian_color_style,
                    subset=numerical_columns,
                    axis=0,
                )
            aggregated_team_season_games.to_excel(
                xlsx_writer,
                float_format="%.2f",
                columns=player_columns,
                header=player_column_names,
                sheet_name=team.name[:31],
            )
            averaged_team_season_games.to_excel(
                xlsx_writer,
                float_format="%.2f",
                columns=player_columns,
                header=player_column_names,
                sheet_name=f"Medias {team.name}"[:31],
            )

    center_alignment = Alignment(horizontal="center")
    for n_sheet, (worksheet_name, worksheet) in enumerate(xlsx_writer.sheets.items()):
        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = center_alignment
        worksheet.column_dimensions[get_column_letter(1)].width = col_width
        worksheet.row_dimensions[1].height = 40
    virtual_workbook = BytesIO()
    xlsx_writer.book.save(virtual_workbook)
    virtual_workbook.seek(0)
    return virtual_workbook.read()
